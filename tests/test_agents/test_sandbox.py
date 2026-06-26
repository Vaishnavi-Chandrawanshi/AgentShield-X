import io
import pytest
import docx
import openpyxl
from backend.app.agents.file_security import FileSecurityAgent

def test_file_scan_no_file():
    agent = FileSecurityAgent()
    result = agent.scan(None)
    assert result.is_safe is True
    assert len(result.detected_threats) == 0
    assert "No file attachment scanned" in result.explanation

def test_file_scan_mz_header():
    agent = FileSecurityAgent()
    file_bytes = b"MZ\x90\x00\x03\x00\x00\x00"
    result = agent.scan(file_bytes, "test.docx")
    assert result.is_safe is False
    assert any("Executable Payload" in t for t in result.detected_threats)

def test_file_scan_mismatch_pdf():
    agent = FileSecurityAgent()
    file_bytes = b"This is not a PDF"
    result = agent.scan(file_bytes, "test.pdf")
    assert result.is_safe is False
    assert any("Mismatched MIME Type" in t for t in result.detected_threats)

def test_file_scan_mismatch_office():
    agent = FileSecurityAgent()
    file_bytes = b"This is not a ZIP file"
    result = agent.scan(file_bytes, "test.docx")
    assert result.is_safe is False
    assert any("Mismatched MIME Type" in t for t in result.detected_threats)

def test_file_scan_yara_office_macro():
    agent = FileSecurityAgent()
    file_bytes = b"PK\x03\x04...Document_Open..."
    result = agent.scan(file_bytes, "test.docx")
    assert result.is_safe is False
    assert any("YARA Alert: OfficeMacroTrigger matched." in t for t in result.detected_threats)

def test_file_scan_yara_pdf_javascript():
    agent = FileSecurityAgent()
    file_bytes = b"%PDF-1.5\n.../JavaScript..."
    result = agent.scan(file_bytes, "test.pdf")
    assert result.is_safe is False
    assert any("YARA Alert: PDFMaliciousTriggers matched." in t for t in result.detected_threats)

def test_file_scan_docx_parsing():
    agent = FileSecurityAgent()
    
    # Generate clean docx
    doc = docx.Document()
    doc.add_paragraph("Hello, world! This is a test paragraph.")
    doc_io = io.BytesIO()
    doc.save(doc_io)
    file_bytes = doc_io.getvalue()
    
    result = agent.scan(file_bytes, "clean.docx")
    assert result.is_safe is True
    assert len(result.detected_threats) == 0
    assert "Hello, world!" in result.extracted_text
    assert "Successful" in result.explanation

def test_file_scan_xlsx_parsing():
    agent = FileSecurityAgent()
    
    # Generate clean xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Column Header"
    ws["A2"] = "Row Value"
    wb_io = io.BytesIO()
    wb.save(wb_io)
    file_bytes = wb_io.getvalue()
    
    result = agent.scan(file_bytes, "clean.xlsx")
    assert result.is_safe is True
    assert len(result.detected_threats) == 0
    assert "Row Value" in result.extracted_text
    assert "Successful" in result.explanation
